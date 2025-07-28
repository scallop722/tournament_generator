#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
トーナメント表自動生成プログラム

ゲーム大会で使用するトーナメント表をExcel形式で出力します。
参加者数に応じて、全組み合わせの対戦表と結果記入表を生成します。
"""

import itertools
from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class TournamentGenerator:
    """トーナメント表生成クラス"""
    
    def __init__(self):
        self.participants = []
        self.matches = []
    
    def generate_participants(self, count: int) -> List[str]:
        """参加者名を生成（A, B, C...）"""
        return [chr(ord('A') + i) for i in range(count)]
    
    def generate_matches(self, participants: List[str]) -> List[Tuple[str, str]]:
        """仕様書の例に基づいた対戦順決定アルゴリズムで対戦カードを生成"""
        # 1. 全組み合わせを生成（並び替え前）
        all_combinations = list(itertools.combinations(participants, 2))
        
        # 2. 参加者毎の対戦数を初期化
        match_count = {participant: 0 for participant in participants}
        
        # 3. 対戦順の決定
        ordered_matches = []
        remaining_combinations = all_combinations.copy()
        
        while remaining_combinations:
            
            # 対戦数が最少の参加者同士の組み合わせを優先探す
            best_match = None
            best_score = float('inf')
            
            for combination in remaining_combinations:
                p1, p2 = combination
                # 両方の参加者の対戦数の合計をスコアとする
                score = match_count[p1] + match_count[p2]
                
                # より少ない対戦数の組み合わせを優先
                if score < best_score:
                    best_score = score
                    best_match = combination
                # 同じスコアの場合、より少ない最小対戦数を持つ組み合わせを優先
                elif score == best_score:
                    current_min = min(match_count[p1], match_count[p2])
                    best_min = min(match_count[best_match[0]], match_count[best_match[1]])
                    if current_min < best_min:
                        best_match = combination
            
            if best_match:
                # 選ばれた組み合わせを結果に追加
                ordered_matches.append(best_match)
                
                # 対戦数を更新
                p1, p2 = best_match
                match_count[p1] += 1
                match_count[p2] += 1
                
                # 使用した組み合わせを削除
                remaining_combinations.remove(best_match)
            else:
                # 万が一のフォールバック：最初の組み合わせを使用
                selected_match = remaining_combinations[0]
                ordered_matches.append(selected_match)
                
                p1, p2 = selected_match
                match_count[p1] += 1
                match_count[p2] += 1
                
                remaining_combinations.remove(selected_match)
        
        # 4. 1P、2P反転アルゴリズムを適用
        balanced_matches = self.apply_player_balance(ordered_matches, participants)
        
        return balanced_matches
    
    def apply_player_balance(self, matches: List[Tuple[str, str]], participants: List[str]) -> List[Tuple[str, str]]:
        """
        1P、2P反転アルゴリズムを適用して公平な1P/2P配置を実現
        
        仕様:
        - 各参加者に重みを設定（初期値0）
        - 1Pになったら+1、2Pになったら-1
        - 対戦時に重みが低い方を1P側に配置
        """
        # 各参加者の重みを初期化
        player_weights = {participant: 0 for participant in participants}
        
        # 各対戦に対して1P/2Pを決定し、重みを更新
        balanced_matches = []
        
        for match in matches:
            p1, p2 = match
            
            # 重みを比較して1P/2Pを決定
            if player_weights[p1] < player_weights[p2]:
                # p1の重みが低いのでp1を1Pに
                first_player = p1
                second_player = p2
            elif player_weights[p1] > player_weights[p2]:
                # p2の重みが低いのでp2を1Pに
                first_player = p2
                second_player = p1
            else:
                # 重みが同じ場合は元の順序を維持
                first_player = p1
                second_player = p2
            
            # 重みを更新（1P: +1, 2P: -1）
            player_weights[first_player] += 1
            player_weights[second_player] -= 1
            
            # 結果に追加
            balanced_matches.append((first_player, second_player))
        
        return balanced_matches
    
    def split_participants(self, participants: List[str]) -> List[List[str]]:
        """参加者を適切なテーブル数に分割"""
        total = len(participants)
        
        if total <= 6:
            # 6人以下は1テーブル
            return [participants]
        elif total <= 12:
            # 7-12人は2テーブル
            mid = total // 2
            return [participants[:mid], participants[mid:]]
        elif total <= 18:
            # 13-18人は3テーブル
            third = total // 3
            remainder = total % 3
            
            tables = []
            start = 0
            for i in range(3):
                # 余りを最初のテーブルに分散
                size = third + (1 if i < remainder else 0)
                tables.append(participants[start:start + size])
                start += size
            
            return tables
        else:
            # 19-24人は4テーブル
            fourth = total // 4
            remainder = total % 4
            
            tables = []
            start = 0
            for i in range(4):
                # 余りを最初のテーブルに分散
                size = fourth + (1 if i < remainder else 0)
                tables.append(participants[start:start + size])
                start += size
            
            return tables
    
    def create_excel_sheet(self, workbook: Workbook, sheet_name: str, 
                          participants: List[str], matches: List[Tuple[str, str]]):
        """Excelシートを作成"""
        ws = workbook.create_sheet(title=sheet_name)
        
        # スタイル設定
        header_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        current_row = 1
        
        # 対戦順テーブル
        ws.cell(row=current_row, column=1, value="対戦順").font = header_font
        current_row += 1
        
        # ヘッダー行
        ws.cell(row=current_row, column=1, value="1P").font = header_font
        ws.cell(row=current_row, column=2, value="-").font = header_font
        ws.cell(row=current_row, column=3, value="2P").font = header_font
        
        # ヘッダー行のスタイル適用
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border
            cell.alignment = center_alignment
        
        current_row += 1
        
        # 対戦カード
        for match in matches:
            ws.cell(row=current_row, column=1, value=match[0])
            ws.cell(row=current_row, column=2, value="vs")
            ws.cell(row=current_row, column=3, value=match[1])
            
            # スタイル適用
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                cell.alignment = center_alignment
            
            current_row += 1
        
        current_row += 2  # 空行
        
        # 結果記入テーブル
        ws.cell(row=current_row, column=1, value="結果記入").font = header_font
        current_row += 1
        
        # 結果記入テーブルのヘッダー
        ws.cell(row=current_row, column=1, value="-").font = header_font
        for i, participant in enumerate(participants):
            ws.cell(row=current_row, column=i + 2, value=participant).font = header_font
        
        # 勝利数、取得本数、順位の列を追加
        wins_col = len(participants) + 2
        points_col = len(participants) + 3
        rank_col = len(participants) + 4
        ws.cell(row=current_row, column=wins_col, value="勝利数").font = header_font
        ws.cell(row=current_row, column=points_col, value="取得本数").font = header_font
        ws.cell(row=current_row, column=rank_col, value="順位").font = header_font
        
        # ヘッダー行のスタイル適用
        for col in range(1, len(participants) + 5):  # +5 = 参加者名 + 参加者列 + 勝利数 + 取得本数 + 順位
            cell = ws.cell(row=current_row, column=col)
            cell.border = border
            cell.alignment = center_alignment
        
        current_row += 1
        
        # 結果記入テーブルの行
        for i, participant in enumerate(participants):
            ws.cell(row=current_row, column=1, value=participant).font = header_font
            
            for j in range(len(participants)):
                if i == j:
                    # 対角線は"-"
                    ws.cell(row=current_row, column=j + 2, value="-")
                else:
                    # 空欄
                    ws.cell(row=current_row, column=j + 2, value="")
                
                # スタイル適用
                cell = ws.cell(row=current_row, column=j + 2)
                cell.border = border
                cell.alignment = center_alignment
            
            # 勝利数、取得本数、順位の列（空欄で初期化）
            wins_cell = ws.cell(row=current_row, column=wins_col, value="")
            points_cell = ws.cell(row=current_row, column=points_col, value="")
            rank_cell = ws.cell(row=current_row, column=rank_col, value="")
            wins_cell.border = border
            wins_cell.alignment = center_alignment
            points_cell.border = border
            points_cell.alignment = center_alignment
            rank_cell.border = border
            rank_cell.alignment = center_alignment
            
            # 参加者名のセルにもスタイル適用
            cell = ws.cell(row=current_row, column=1)
            cell.border = border
            cell.alignment = center_alignment
            
            current_row += 1
        
        # 列幅調整
        for col in range(1, len(participants) + 5):  # +5 = 参加者名 + 参加者列 + 勝利数 + 取得本数 + 順位
            column_letter = get_column_letter(col)
            if col == wins_col or col == points_col or col == rank_col:
                # 勝利数・取得本数・順位列は少し幅を広く
                ws.column_dimensions[column_letter].width = 10
            else:
                ws.column_dimensions[column_letter].width = 8
    
    def generate_tournament(self, participant_count: int) -> str:
        """トーナメント表を生成してExcelファイルを作成"""
        # 参加者生成
        all_participants = self.generate_participants(participant_count)
        
        # 参加者をテーブルに分割
        tables = self.split_participants(all_participants)
        
        # Excelワークブック作成
        workbook = Workbook()
        # デフォルトシートを削除
        workbook.remove(workbook.active)
        
        # 各テーブルのシートを作成
        for i, table_participants in enumerate(tables):
            # 対戦カード生成
            matches = self.generate_matches(table_participants)
            
            # シート名決定
            if len(tables) == 1:
                sheet_name = "トーナメント表"
            else:
                sheet_name = f"テーブル{chr(ord('A') + i)}"
            
            # シート作成
            self.create_excel_sheet(workbook, sheet_name, table_participants, matches)
        
        # ファイル名生成
        filename = f"tournament_{participant_count}人.xlsx"
        
        # ファイル保存
        workbook.save(filename)
        
        return filename


def main():
    """メイン関数"""
    print("=== トーナメント表自動生成プログラム ===")
    print("参加者数を入力してください（3～24人）")
    
    try:
        participant_count = int(input("参加者数: "))
        
        if participant_count < 3:
            print("エラー: 参加者数は3人以上で入力してください。")
            return
        
        if participant_count > 24:
            print("エラー: 参加者数は24人以下で入力してください。")
            return
        
        # トーナメント生成
        generator = TournamentGenerator()
        filename = generator.generate_tournament(participant_count)
        
        print(f"トーナメント表を生成しました: {filename}")
        
        # 生成された内容の概要を表示
        all_participants = generator.generate_participants(participant_count)
        tables = generator.split_participants(all_participants)
        
        print(f"\n=== 生成内容 ===")
        print(f"参加者数: {participant_count}人")
        print(f"テーブル数: {len(tables)}個")
        
        for i, table in enumerate(tables):
            matches = generator.generate_matches(table)
            print(f"テーブル{chr(ord('A') + i) if len(tables) > 1 else ''}: {len(table)}人, {len(matches)}試合")
            print(f"  参加者: {', '.join(table)}")
        
    except ValueError:
        print("エラー: 数値を入力してください。")
    except Exception as e:
        print(f"エラーが発生しました: {e}")


if __name__ == "__main__":
    main()
